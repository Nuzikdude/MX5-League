from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any, Dict

from openpyxl import load_workbook


POINTS_MAP = {1: 10, 2: 7, 3: 5, 4: 3, 5: 2, 6: 1}


def slugify(value: str) -> str:
    value = str(value or "").strip().lower()
    value = value.replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value


def clean_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_int(value: Any):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        try:
            return int(float(value))
        except Exception:
            return None


def read_json_if_exists(path: Path, default):
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path: Path, data: Any):
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def build_export(workbook_path: Path, data_dir: Path):
    wb = load_workbook(workbook_path, data_only=True, keep_vba=True)

    overview = wb["League Overview"]

    existing_drivers = read_json_if_exists(data_dir / "drivers.json", [])
    existing_tracks = read_json_if_exists(data_dir / "tracks.json", [])

    existing_driver_by_tab = {d.get("tab"): d for d in existing_drivers}
    existing_driver_by_id = {d.get("id"): d for d in existing_drivers}
    existing_track_by_id = {t.get("id"): t for t in existing_tracks}
    existing_track_by_name = {t.get("name"): t for t in existing_tracks}

    league_name = clean_str(overview["A1"].value) or "MX5 League"

    # -------------------------
    # RULES EXPORT
    # -------------------------
    rules = []
    if "Rules" in wb.sheetnames:
        rules_ws = wb["Rules"]
        for row in rules_ws.iter_rows(min_row=1, max_col=2, values_only=True):
            label, value = row
            label = clean_str(label)
            value = clean_str(value)
            if label or value:
                rules.append({"label": label, "value": value})
    else:
        for row in range(4, 12):
            label = clean_str(overview[f"A{row}"].value)
            value = clean_str(overview[f"B{row}"].value)
            if label or value:
                rules.append({"label": label, "value": value})

    league = {
        "name": league_name,
        "rules": rules,
    }

    # -------------------------
    # DRIVERS
    # -------------------------
    drivers = []
    driver_pages = []

    for row in range(4, 10):
        tab = clean_str(overview[f"D{row}"].value)
        listed_name = clean_str(overview[f"E{row}"].value)

        if not tab:
            continue

        driver_ws = wb[tab] if tab in wb.sheetnames else None

        name = listed_name
        nationality = ""
        nickname = ""
        age = None

        if driver_ws is not None:
            name = clean_str(driver_ws["B1"].value) or name
            nationality = clean_str(driver_ws["B2"].value)
            nickname = clean_str(driver_ws["B3"].value)
            age = clean_int(driver_ws["B4"].value)

        driver_idx = row - 3
        driver_id = f"driver-{driver_idx}"

        existing = existing_driver_by_tab.get(tab) or existing_driver_by_id.get(driver_id) or {}
        image = existing.get("image") or f"/images/drivers/driver{driver_idx}.jpg"

        drivers.append({
            "id": driver_id,
            "tab": tab,
            "name": name,
            "nationality": nationality,
            "nickname": nickname,
            "age": age,
            "image": image,
        })

        driver_pages.append(drivers[-1].copy())

    driver_names = [d["name"] for d in drivers]

    # -------------------------
    # TRACKS
    # -------------------------
    tracks = []
    track_pages = []

    for row in range(13, 32):
        raw_track_name = clean_str(overview[f"A{row}"].value)
        laps = clean_int(overview[f"B{row}"].value)

        if not raw_track_name or raw_track_name.lower() == "track":
            continue

        possible_sheetnames = [
            raw_track_name,
            raw_track_name.replace(" ", "_"),
            raw_track_name.replace("-", "_"),
            raw_track_name.replace(" ", "-"),
        ]
        sheet_name = next((s for s in possible_sheetnames if s in wb.sheetnames), None)

        track_id = slugify(raw_track_name)
        existing_track = existing_track_by_id.get(track_id) or existing_track_by_name.get(raw_track_name) or {}

        image = existing_track.get("image") or f"/images/tracks/{track_id}.png"
        location = existing_track.get("location", "")

        link = ""
        results = []

        if sheet_name:
            ws = wb[sheet_name]
            page_name = clean_str(ws["A1"].value) or raw_track_name
            link = clean_str(ws["P1"].value)

            if not laps:
                laps = clean_int(existing_track.get("laps"))

            for result_row in range(4, 10):
                position = clean_str(ws[f"A{result_row}"].value)
                driver_name = clean_str(ws[f"B{result_row}"].value)
                race_points = clean_int(ws[f"C{result_row}"].value)
                total_points = clean_int(ws[f"E{result_row}"].value)
                lap_time = clean_str(ws[f"F{result_row}"].value)
                fastest_raw = clean_str(ws[f"D{result_row}"].value).upper()

                has_real_result = (
                    lap_time not in {"", "-", "0", "0:00", "00:00", "00:00.000"}
                    or fastest_raw in {"Y", "YES", "TRUE", "1", "🔥"}
                )

                if not driver_name and result_row - 4 < len(driver_names):
                    driver_name = driver_names[result_row - 4]

                results.append({
                    "position": position if has_real_result else "",
                    "driver": driver_name,
                    "racePoints": race_points if race_points is not None else "",
                    "fastestLap": fastest_raw in {"Y", "YES", "TRUE", "1", "🔥"},
                    "totalPoints": total_points if total_points is not None else "",
                    "lapTime": lap_time,
                    "bestLapTime": lap_time,
                    "completed": has_real_result,
                })

            track_name = page_name
        else:
            track_name = raw_track_name
            for driver_name in driver_names:
                results.append({
                    "position": "",
                    "driver": driver_name,
                    "racePoints": "",
                    "fastestLap": False,
                    "totalPoints": "",
                    "lapTime": "",
                    "bestLapTime": "",
                    "completed": False,
                })

        track_summary = {
            "id": track_id,
            "name": track_name,
            "laps": laps,
            "image": image,
            "tab": sheet_name or raw_track_name,
        }

        if link:
            track_summary["link"] = link
        if location:
            track_summary["location"] = location

        track_page = dict(track_summary)
        track_page["results"] = results

        tracks.append(track_summary)
        track_pages.append(track_page)

    # -------------------------
    # STANDINGS
    # -------------------------
    standings_table: Dict[str, Dict[str, Any]] = {}
    finish_sums: Dict[str, int] = {}
    finish_counts: Dict[str, int] = {}

    for driver in drivers:
        standings_table[driver["name"]] = {
            "driver": driver["name"],
            "points": 0,
            "wins": 0,
            "podiums": 0,
            "fastLaps": 0,
            "starts": 0,
            "avgFinish": "",
        }

    for track in track_pages:
        for r in track.get("results", []):
            name = clean_str(r.get("driver"))
            if not name or name not in driver_names:
                continue

            if not bool(r.get("completed")):
                continue

            pos = clean_int(r.get("position"))
            if pos is None or pos < 1 or pos > 6:
                continue

            standings_table[name]["starts"] += 1
            standings_table[name]["points"] += POINTS_MAP.get(pos, 0)

            if pos == 1:
                standings_table[name]["wins"] += 1
            if pos <= 3:
                standings_table[name]["podiums"] += 1

            finish_sums[name] = finish_sums.get(name, 0) + pos
            finish_counts[name] = finish_counts.get(name, 0) + 1

            if bool(r.get("fastestLap")):
                standings_table[name]["fastLaps"] += 1
                standings_table[name]["points"] += 2

    standings = list(standings_table.values())

    for entry in standings:
        name = entry["driver"]
        if finish_counts.get(name):
            entry["avgFinish"] = round(finish_sums[name] / finish_counts[name], 2)

    standings.sort(
        key=lambda x: (
            -int(x.get("points", 0) or 0),
            -int(x.get("wins", 0) or 0),
            -int(x.get("podiums", 0) or 0),
            x.get("driver", ""),
        )
    )

    for i, entry in enumerate(standings, start=1):
        entry["rank"] = i

    # -------------------------
    # NEWS
    # -------------------------
    news = []

    if "News" in wb.sheetnames:
        news_ws = wb["News"]

        for row in (1, 13, 25, 37, 49, 61, 73, 85):
            raw_date = news_ws[f"B{row}"].value
            headline = clean_str(news_ws[f"B{row + 1}"].value)
            body = clean_str(news_ws[f"A{row + 2}"].value)

            if not headline and not body:
                continue

            date = raw_date.strftime("%d %B %Y") if hasattr(raw_date, "strftime") else clean_str(raw_date)

            news.append({
                "date": date,
                "headline": headline,
                "body": body,
            })

    # -------------------------
    # CALENDAR (FIXED COMPLETED SUPPORT)
    # -------------------------
    calendar = []

    for row in range(13, 32):
        track = clean_str(overview[f"A{row}"].value)
        laps = clean_int(overview[f"B{row}"].value)
        date = clean_str(overview[f"D{row}"].value)
        time = clean_str(overview[f"E{row}"].value)
        
        # Get the raw value from Column C (Completed column)
        completed_raw = str(overview[f"C{row}"].value or "").strip().lower()

        if not track or track.lower() == "track":
            continue

        # 1. Manual check: Add "x" or checkmark support
        is_marked_completed = completed_raw in {"y", "yes", "true", "1", "x", "completed"}

        # 2. Automatic check: See if this track already has processed results
        track_id = slugify(track)
        has_results = False
        
        # Look through the track_pages we built earlier to see if any results are 'completed'
        track_info = next((tp for tp in track_pages if tp["id"] == track_id), None)
        if track_info:
            has_results = any(r.get("completed") for r in track_info.get("results", []))

        calendar.append({
            "track": track,
            "laps": laps,
            "date": date,
            "time": time,
            "completed": is_marked_completed or has_results, # True if either condition is met
        })

    # -------------------------
    # SAVE OUTPUTS
    # -------------------------
    data_dir.mkdir(parents=True, exist_ok=True)

    save_json(data_dir / "league.json", league)
    save_json(data_dir / "drivers.json", drivers)
    save_json(data_dir / "driver-pages.json", driver_pages)
    save_json(data_dir / "tracks.json", tracks)
    save_json(data_dir / "track-pages.json", track_pages)
    save_json(data_dir / "standings.json", standings)
    save_json(data_dir / "news.json", news)
    save_json(data_dir / "calendar.json", calendar)
    save_json(data_dir / "rules.json", rules)

    print("Export complete.")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python export_mx5_json.py <path-to-xlsm> <path-to-public-data-folder>")
        sys.exit(1)

    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    data_dir = Path(sys.argv[2]).expanduser().resolve()

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        sys.exit(1)

    build_export(workbook_path, data_dir)